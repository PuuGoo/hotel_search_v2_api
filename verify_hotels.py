import argparse
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from typing import Dict, Tuple
from urllib.parse import urlparse, urlunparse

import pandas as pd
from rapidfuzz import fuzz
from unidecode import unidecode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except Exception:
    webdriver = None
    By = None
    Service = None
    WebDriverWait = None
    ChromeDriverManager = None
    SELENIUM_AVAILABLE = False


BLOCKED_OR_ERROR_PATTERNS = [
    "this site cant be reached",
    "this site can t be reached",
    "err_http2_protocol_error",
    "403 forbidden",
    "410 gone",
    "performing security verification",
    "lets confirm you are human",
    "are you a person or a robot",
    "complete the security check",
    "access denied",
    "captcha",
    "cloudflare",
]

GENERIC_NAME_WORDS = {
    "hotel",
    "hotels",
    "hostel",
    "guesthouse",
    "guest",
    "house",
    "resort",
    "inn",
    "apartments",
    "apartment",
}


def normalize_text(value: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = unidecode(str(value)).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return " ".join(text.split())


def normalize_name_for_matching(value: str) -> str:
    tokens = [t for t in normalize_text(value).split() if t and t not in GENERIC_NAME_WORDS]
    return " ".join(tokens)


def compute_name_score(expected_name: str, candidate_name: str) -> float:
    score_raw = fuzz.token_set_ratio(expected_name, candidate_name)
    score_generic_removed = fuzz.token_set_ratio(
        normalize_name_for_matching(expected_name),
        normalize_name_for_matching(candidate_name),
    )
    return max(score_raw, score_generic_removed)


def extract_number_tokens(value: str):
    return set(re.findall(r"\d+[a-z]?", value or ""))


def addresses_semantically_match(master_address: str, child_address: str) -> Tuple[bool, str, float]:
    if not master_address or not child_address:
        return False, "address_missing", 0

    if master_address == child_address:
        return True, "address_exact", 100

    token_set_score = fuzz.token_set_ratio(master_address, child_address)
    partial_score = fuzz.partial_ratio(master_address, child_address)
    best_score = max(token_set_score, partial_score)

    if best_score >= 70:
        return True, "address_fuzzy_high", best_score

    shorter = master_address if len(master_address) <= len(child_address) else child_address
    longer = child_address if shorter == master_address else master_address
    if len(shorter) >= 8 and shorter in longer:
        return True, "address_contains", best_score

    master_tokens = {t for t in master_address.split() if len(t) > 2}
    child_tokens = {t for t in child_address.split() if len(t) > 2}
    common_tokens = master_tokens.intersection(child_tokens)
    if master_tokens and child_tokens:
        token_overlap_ratio = len(common_tokens) / min(len(master_tokens), len(child_tokens))
        if len(common_tokens) >= 2 and token_overlap_ratio >= 0.5 and best_score >= 65:
            return True, "address_token_overlap", best_score

    master_numbers = extract_number_tokens(master_address)
    child_numbers = extract_number_tokens(child_address)
    if master_numbers and child_numbers and master_numbers.intersection(child_numbers) and best_score >= 55:
        return True, "address_number_overlap", best_score

    return False, "address_low_similarity", best_score


def infer_name_from_url(url: str) -> str:
    try:
        parsed = urlparse(url)
        segments = [s for s in parsed.path.split("/") if s]
        for seg in reversed(segments):
            seg = seg.strip().lower()
            if not seg:
                continue
            seg = re.sub(r"\.(html|htm)$", "", seg)
            words = [w for w in re.split(r"[-_]", seg) if re.search(r"[a-z]", w)]
            if len(words) >= 3:
                return " ".join(words)
    except Exception:
        pass
    return ""


def get_url_candidates(url: str):
    candidates = [url]
    try:
        parsed = urlparse(url)
        clean = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
        if clean and clean not in candidates:
            candidates.append(clean)

        host = (parsed.netloc or "").lower()
        if host.endswith(".hotels.com") and host != "www.hotels.com":
            fallback_host = "www.hotels.com"
            fallback_url = urlunparse((parsed.scheme or "https", fallback_host, parsed.path, "", "", ""))
            if fallback_url not in candidates:
                candidates.append(fallback_url)
    except Exception:
        pass
    return candidates


def load_page_with_fallback(driver, url: str):
    last_error = ""
    candidates = get_url_candidates(url)

    for candidate in candidates:
        for _ in range(1):
            try:
                driver.get(candidate)
                WebDriverWait(driver, 8).until(
                    lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")
                )
                time.sleep(0.2)

                page_title = driver.title or ""
                try:
                    body_text = driver.find_element(By.TAG_NAME, "body").text[:1200]
                except Exception:
                    body_text = ""

                if is_blocked_or_error_page(page_title, body_text):
                    last_error = f"blocked_or_error_page on {candidate}"
                    continue

                return
            except Exception as ex:
                last_error = str(ex)
                time.sleep(0.2)

    if last_error:
        raise RuntimeError(last_error)
    raise RuntimeError("cannot_load_page")


def extract_address_from_json_ld(driver) -> str:
    scripts = driver.find_elements(By.CSS_SELECTOR, 'script[type="application/ld+json"]')

    def walk(node):
        if isinstance(node, dict):
            address = node.get("address")
            if isinstance(address, str) and address.strip():
                return address.strip()
            if isinstance(address, dict):
                parts = [
                    address.get("streetAddress", ""),
                    address.get("addressLocality", ""),
                    address.get("addressRegion", ""),
                    address.get("postalCode", ""),
                    address.get("addressCountry", ""),
                ]
                merged = ", ".join([p for p in parts if p])
                if merged:
                    return merged

            for value in node.values():
                result = walk(value)
                if result:
                    return result
        elif isinstance(node, list):
            for item in node:
                result = walk(item)
                if result:
                    return result
        return ""

    for script in scripts:
        raw = (script.get_attribute("textContent") or "").strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
            result = walk(data)
            if result:
                return result
        except Exception:
            continue

    return ""


def is_blocked_or_error_page(found_name: str, found_address: str) -> bool:
    combined = normalize_text(f"{found_name} {found_address}")
    return any(p in combined for p in BLOCKED_OR_ERROR_PATTERNS)


def text_contains_phrase(text: str, phrase: str) -> bool:
    if not text or not phrase:
        return False


def resolve_data_file_path(file_name: str) -> str:
    if os.path.exists(file_name):
        return file_name

    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(base_dir, file_name)
    if os.path.exists(candidate):
        return candidate

    return file_name
    if f" {phrase} " in f" {text} ":
        return True

    compact_text = re.sub(r"[^a-z0-9]", "", text)
    compact_phrase = re.sub(r"[^a-z0-9]", "", phrase)
    if len(compact_phrase) >= 6 and compact_phrase in compact_text:
        return True

    phrase_tokens = [t for t in phrase.split() if t]
    if len(phrase_tokens) >= 2:
        text_tokens = set(text.split())
        if all(token in text_tokens for token in phrase_tokens):
            return True

    return False


def is_valid_brand_phrase(value: str) -> bool:
    normalized = normalize_text(value)
    if not normalized:
        return False
    tokens = normalized.split()
    if len(normalized) < 2:
        return False
    if normalized in {"as", "the", "and", "for", "by", "in", "on", "of", "to"}:
        return False
    if len(tokens) == 1 and tokens[0] in GENERIC_NAME_WORDS:
        return False
    return True


def split_brand_aliases(value: str):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []

    raw_text = unidecode(str(value)).lower()
    parts = re.split(r"\s*/\s*|\s*,\s*|\s*;\s*|\s*\|\s*|\s*--\s*", raw_text)
    aliases = []
    seen = set()
    for part in parts:
        candidate_variants = [part, re.sub(r"\(.*?\)", "", part)]
        for variant in candidate_variants:
            candidate = normalize_text(variant)
            if not candidate or candidate in seen:
                continue
            if not is_valid_brand_phrase(candidate):
                continue
            seen.add(candidate)
            aliases.append(candidate)
    return aliases


def load_chain_subbrand_data(chain_file_path: str):
    chain_file_path = resolve_data_file_path(chain_file_path)
    if not os.path.exists(chain_file_path):
        return [], [], []

    df_chain = pd.read_excel(chain_file_path)
    if len(df_chain.columns) < 2:
        return [], [], []

    col0, col1 = df_chain.columns[0], df_chain.columns[1]
    temp = df_chain[[col0, col1]].copy()
    temp.columns = ["chain", "sub_brand"]
    temp["chain"] = temp["chain"].ffill()

    chain_list = []
    chain_subbrand_pairs = []
    chain_alias_groups = []
    seen_chain = set()
    seen_pairs = set()

    for _, row in temp.iterrows():
        chain = normalize_text(row["chain"])
        sub_brand = normalize_text(row["sub_brand"])

        if not chain or chain in {"chain", "danh sach sub brand tuong ung cua chain"}:
            continue
        if sub_brand in {"sub brand", "sub brand", "subbrand"}:
            sub_brand = ""

        chain_aliases = split_brand_aliases(row["chain"])
        sub_brand_aliases = split_brand_aliases(row["sub_brand"])

        if len(chain_aliases) > 1:
            chain_alias_groups.append(set(chain_aliases))

        for chain_alias in chain_aliases:
            if chain_alias not in seen_chain:
                seen_chain.add(chain_alias)
                chain_list.append(chain_alias)

        if sub_brand_aliases:
            for chain_alias in chain_aliases:
                for sub_alias in sub_brand_aliases:
                    key = (chain_alias, sub_alias)
                    if key not in seen_pairs:
                        seen_pairs.add(key)
                        chain_subbrand_pairs.append(key)

    return chain_list, chain_subbrand_pairs, chain_alias_groups


def load_vho_terms(vho_file_path: str):
    vho_file_path = resolve_data_file_path(vho_file_path)
    if not os.path.exists(vho_file_path):
        return []

    df_vho = pd.read_excel(vho_file_path)
    if len(df_vho.columns) < 2:
        return []

    col0, col1 = df_vho.columns[0], df_vho.columns[1]
    temp = df_vho[[col0, col1]].copy()
    temp.columns = ["vho", "sub_brand"]
    temp["vho"] = temp["vho"].ffill()

    terms = []
    seen = set()

    for _, row in temp.iterrows():
        for alias in split_brand_aliases(row["vho"]):
            if alias in {"vho", "danh sach vho", "sub brand", "subbrand"}:
                continue
            if alias not in seen:
                seen.add(alias)
                terms.append(alias)
        for alias in split_brand_aliases(row["sub_brand"]):
            if alias in {"vho", "danh sach vho", "sub brand", "subbrand"}:
                continue
            if alias not in seen:
                seen.add(alias)
                terms.append(alias)

    if "captial o" in seen and "capital o" not in seen:
        seen.add("capital o")
        terms.append("capital o")

    if "oyo" in seen and "hotel o" not in seen:
        seen.add("hotel o")
        terms.append("hotel o")

    return sorted(terms, key=len, reverse=True)


def highlight_column_yellow(excel_path: str, column_name: str):
    wb = load_workbook(excel_path)
    ws = wb.active

    target_col = None
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value or "").strip() == column_name:
            target_col = col_idx
            break

    if target_col is None:
        wb.close()
        return

    yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    for row_idx in range(1, ws.max_row + 1):
        ws.cell(row=row_idx, column=target_col).fill = yellow_fill

    wb.save(excel_path)
    wb.close()


def classify_chain_branch_case(master_name: str, child_name: str, chain_list, chain_subbrand_pairs, chain_alias_groups, vho_terms):
    master_vho_terms = [term for term in vho_terms if text_contains_phrase(master_name, term)]
    child_vho_terms = [term for term in vho_terms if text_contains_phrase(child_name, term)]
    if master_vho_terms and child_vho_terms:
        return f"Cả master và child đều chứa VHO: master={master_vho_terms[0]}; child={child_vho_terms[0]}"
    if master_vho_terms and not child_vho_terms:
        return f"Chỉ master chứa VHO: master={master_vho_terms[0]}"
    if child_vho_terms and not master_vho_terms:
        return f"Chỉ child chứa VHO: child={child_vho_terms[0]}"

    if master_name and child_name and master_name == child_name:
        return ""

    branch_terms = set(chain_list)
    branch_terms.update(sub_brand for _, sub_brand in chain_subbrand_pairs)
    ordered_branch_terms = sorted(branch_terms, key=len, reverse=True)

    for alias_group in chain_alias_groups:
        master_has_group = any(text_contains_phrase(master_name, alias) for alias in alias_group)
        child_has_group = any(text_contains_phrase(child_name, alias) for alias in alias_group)
        if master_has_group and child_has_group:
            return ""

    master_terms = [term for term in ordered_branch_terms if text_contains_phrase(master_name, term)]
    child_terms = [term for term in ordered_branch_terms if text_contains_phrase(child_name, term)]

    if master_terms and child_terms:
        if not set(master_terms).intersection(set(child_terms)):
            return f"Chain/branch khác nhau: master={master_terms[0]}; child={child_terms[0]}"

    for chain in ordered_branch_terms:
        master_has_chain = text_contains_phrase(master_name, chain)
        child_has_chain = text_contains_phrase(child_name, chain)
        if master_has_chain and child_has_chain:
            return ""

    for chain, sub_brand in chain_subbrand_pairs:
        master_has_chain = text_contains_phrase(master_name, chain)
        child_has_chain = text_contains_phrase(child_name, chain)
        master_has_sub_brand = text_contains_phrase(master_name, sub_brand)
        child_has_sub_brand = text_contains_phrase(child_name, sub_brand)

        if master_has_chain and child_has_sub_brand:
            return f"Chain/branch trên master chứa sub-brand child: chain={chain}; sub-brand={sub_brand}"
        if child_has_chain and master_has_sub_brand:
            return f"Chain/branch trên child chứa sub-brand master: chain={chain}; sub-brand={sub_brand}"

    for chain in ordered_branch_terms:
        master_has_chain = text_contains_phrase(master_name, chain)
        child_has_chain = text_contains_phrase(child_name, chain)
        if master_has_chain and not child_has_chain:
            return f"Chain/branch trên master: chain={chain}"

    return ""


def get_input_columns(df: pd.DataFrame) -> Dict[str, str | bool]:
    cols = list(df.columns)
    if len(cols) < 3:
        raise ValueError("File Excel phải có ít nhất 3 cột: tên khách sạn, địa chỉ, link OTA.")

    normalized = {c: normalize_text(c) for c in cols}

    def find_col(keywords):
        for c, n in normalized.items():
            if any(k in n for k in keywords):
                return c
        return None

    def find_col_with_all(parts):
        for c, n in normalized.items():
            if all(p in n for p in parts):
                return c
        return None

    found_master_name_col = find_col_with_all(["master", "hotel", "name"])
    found_child_name_col = find_col_with_all(["child", "hotel", "name"])
    found_master_address_col = find_col_with_all(["master", "hotel", "address"])
    found_child_address_col = find_col_with_all(["child", "hotel", "address"])
    ota_link_col = find_col_with_all(["url", "ota", "child"])

    hotel_name_col = find_col(["hotel", "khach san", "ten"])
    address_col = find_col(["address", "dia chi"])
    if not ota_link_col:
        ota_link_col = find_col(["ota", "link", "url"])

    child_name_col = found_child_name_col or hotel_name_col or cols[0]
    child_address_col = found_child_address_col or address_col or cols[1]

    master_name_col = found_master_name_col or child_name_col
    master_address_col = found_master_address_col or child_address_col

    if not ota_link_col:
        ota_link_col = cols[2]

    has_master_child_layout = bool(
        found_master_name_col and found_child_name_col and found_master_address_col and found_child_address_col
    )

    return {
        "master_name_col": master_name_col,
        "child_name_col": child_name_col,
        "master_address_col": master_address_col,
        "child_address_col": child_address_col,
        "ota_link_col": ota_link_col,
        "has_master_child_layout": has_master_child_layout,
    }


def first_text(driver, selectors, timeout_seconds: float = 2.0):
    end_time = time.time() + timeout_seconds
    while time.time() < end_time:
        for by, selector in selectors:
            try:
                elements = driver.find_elements(by, selector)
                if not elements:
                    continue
                element = elements[0]
                text = (element.text or element.get_attribute("content") or "").strip()
                if text:
                    return text
            except Exception:
                continue
        time.sleep(0.1)
    return ""


def extract_page_data(driver, url: str) -> Tuple[str, str]:
    load_page_with_fallback(driver, url)

    domain = ""
    try:
        domain = url.split("/")[2].lower()
    except Exception:
        pass

    common_name_selectors = [
        (By.CSS_SELECTOR, "h1"),
        (By.CSS_SELECTOR, "h2"),
        (By.CSS_SELECTOR, 'meta[property="og:title"]'),
        (By.CSS_SELECTOR, "title"),
    ]
    common_address_selectors = [
        (By.CSS_SELECTOR, '[itemprop="streetAddress"]'),
        (By.CSS_SELECTOR, '[data-testid="address"]'),
        (By.CSS_SELECTOR, '[data-testid*="address"]'),
        (By.CSS_SELECTOR, '[class*="Address"]'),
        (By.CSS_SELECTOR, '[class*="address"]'),
    ]

    fallback_address_selectors = [
        (By.CSS_SELECTOR, 'meta[property="hotel:address"]'),
        (By.CSS_SELECTOR, 'meta[name="description"]'),
        (By.CSS_SELECTOR, 'meta[property="og:description"]'),
    ]

    per_domain = {
        "booking.com": {
            "name": [(By.CSS_SELECTOR, 'h2[data-testid="title"]')],
            "address": [(By.CSS_SELECTOR, 'span[data-testid="address"]')],
        },
        "agoda.com": {
            "name": [(By.CSS_SELECTOR, '[data-selenium="hotel-header-name"]')],
            "address": [(By.CSS_SELECTOR, '[data-selenium="hotel-address-map"]')],
        },
        "traveloka.com": {
            "name": [(By.CSS_SELECTOR, "h1")],
            "address": [(By.CSS_SELECTOR, '[data-testid*="Address"]')],
        },
    }

    name_selectors = []
    address_selectors = []

    for d, selectors in per_domain.items():
        if d in domain:
            name_selectors.extend(selectors["name"])
            address_selectors.extend(selectors["address"])
            break

    name_selectors.extend(common_name_selectors)
    address_selectors.extend(common_address_selectors)

    found_name = first_text(driver, name_selectors) or (driver.title or "")
    found_address = first_text(driver, address_selectors)

    if not found_address:
        found_address = extract_address_from_json_ld(driver)

    if not found_address:
        found_address = first_text(driver, fallback_address_selectors)

    if not found_address:
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            found_address = body_text[:3000]
        except Exception:

            found_address = ""

    return found_name, found_address


def create_driver(headless: bool):
    if not SELENIUM_AVAILABLE:
        raise RuntimeError("Selenium/Chrome dependencies are not available in this runtime")

    options = webdriver.ChromeOptions()
    options.page_load_strategy = "eager"
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--blink-settings=imagesEnabled=false")
    options.add_experimental_option(
        "prefs",
        {
            "profile.managed_default_content_settings.images": 2,
            "profile.managed_default_content_settings.stylesheets": 2,
            "profile.managed_default_content_settings.fonts": 2,
        },
    )
    if headless:
        options.add_argument("--headless=new")
    else:
        options.add_argument("--window-size=800,600")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.set_page_load_timeout(35)
    return driver


def process_single_row(
    row_data,
    master_name_col,
    child_name_col,
    master_address_col,
    child_address_col,
    ota_link_col,
    has_master_child_layout,
    name_threshold,
    address_threshold,
    strong_name_threshold,
    url_name_threshold,
    headless,
    page_cache,
    cache_lock,
):
    index, row = row_data
    master_name = normalize_text(row[master_name_col])
    child_name = normalize_text(row[child_name_col])
    master_address = normalize_text(row[master_address_col])
    child_address = normalize_text(row[child_address_col])

    name_score_master_child = fuzz.ratio(master_name, child_name)
    name_exact_match = bool(master_name and child_name and master_name == child_name)
    address_match, address_match_mode, address_match_score = addresses_semantically_match(master_address, child_address)

    if has_master_child_layout and name_exact_match and address_match:
        return {
            "index": index,
            "matched": True,
            "found_name": child_name,
            "found_address": child_address,
            "name_score": name_score_master_child,
            "address_score": address_match_score,
            "match_reason": "master_child_precheck_match",
            "check_url_score": 0,
            "check_url_pass": False,
            "check_link": False,
            "action_note": f"Match pre-check: tên + địa chỉ master/child ({address_match_mode})",
        }

    expected_name = child_name
    expected_address = child_address
    ota_url = str(row[ota_link_col]).strip()

    is_match = False
    found_name = ""
    found_address = ""
    name_score = 0
    address_score = 0
    match_reason = "invalid_url"
    check_url_score = 0
    check_url_pass = False
    check_link = False
    action_note = "Bỏ qua: URL không hợp lệ"

    if ota_url.startswith("http"):
        inferred_name = infer_name_from_url(ota_url)
        url_name_score = compute_name_score(expected_name, inferred_name) if inferred_name else 0
        check_url_score = url_name_score
        check_url_pass = url_name_score >= url_name_threshold

        if url_name_score >= url_name_threshold:
            return {
                "index": index,
                "matched": True,
                "found_name": inferred_name,
                "found_address": "",
                "name_score": url_name_score,
                "address_score": 0,
                "match_reason": "url_name_high_confidence",
                "check_url_score": check_url_score,
                "check_url_pass": check_url_pass,
                "check_link": check_link,
                "action_note": "Match: tên khớp cao từ URL, bỏ qua crawl",
            }

        with cache_lock:
            cached = page_cache.get(ota_url)

        if cached:
            if cached.get("error"):
                found_name = cached.get("found_name", "")
                name_score = cached.get("name_score", 0)
                match_reason = "processing_exception"
            else:
                found_name = cached.get("found_name", "")
                found_address = cached.get("found_address", "")
        else:
            driver = None
            try:
                check_link = True
                driver = create_driver(headless)
                found_name, found_address = extract_page_data(driver, ota_url)
                with cache_lock:
                    page_cache[ota_url] = {
                        "found_name": found_name,
                        "found_address": found_address,
                        "error": "",
                    }
            except Exception as ex:
                inferred_name = infer_name_from_url(ota_url)
                if inferred_name:
                    found_name = inferred_name
                    name_score = compute_name_score(expected_name, inferred_name)

                with cache_lock:
                    page_cache[ota_url] = {
                        "found_name": found_name,
                        "found_address": "",
                        "error": str(ex),
                        "name_score": name_score,
                    }

                if "blocked_or_error_page" in normalize_text(str(ex)):
                    match_reason = "blocked_or_error_page"
                    action_note = "Crawl lỗi: bị chặn hoặc trang lỗi"
                else:
                    match_reason = "processing_exception"
                    action_note = "Crawl lỗi: exception khi xử lý"

                return {
                    "index": index,
                    "matched": False,
                    "found_name": found_name,
                    "found_address": "",
                    "name_score": name_score,
                    "address_score": 0,
                    "match_reason": match_reason,
                    "check_url_score": check_url_score,
                    "check_url_pass": check_url_pass,
                    "check_link": check_link,
                    "action_note": action_note,
                }

            finally:
                if driver:
                    try:
                        driver.quit()
                    except Exception:
                        pass

        normalized_found_name = normalize_text(found_name)
        normalized_found_address = normalize_text(found_address)

        inferred_name = infer_name_from_url(ota_url)
        name_score_page = compute_name_score(expected_name, normalized_found_name)
        name_score_url = compute_name_score(expected_name, inferred_name) if inferred_name else 0

        if name_score_url > name_score_page and name_score_url >= 80:
            name_score = name_score_url
            found_name = inferred_name
        else:
            name_score = name_score_page

        address_score = fuzz.token_set_ratio(expected_address, normalized_found_address)

        blocked_or_error = is_blocked_or_error_page(found_name, found_address)
        if blocked_or_error:
            is_match = False
            match_reason = "blocked_or_error_page"
            action_note = "Crawl xong: trang bị chặn/lỗi"
        elif name_score >= name_threshold and address_score >= address_threshold:
            is_match = True
            match_reason = "name_and_address_matched"
            action_note = "Crawl xong: khớp tên + địa chỉ"
        elif name_score >= strong_name_threshold:
            is_match = True
            match_reason = "strong_name_match_fallback"
            action_note = "Crawl xong: khớp mạnh theo tên"
        else:
            is_match = False
            match_reason = "low_similarity"
            action_note = "Crawl xong: độ tương đồng thấp"

    return {
        "index": index,
        "matched": bool(is_match),
        "found_name": found_name,
        "found_address": found_address,
        "name_score": name_score,
        "address_score": address_score,
        "match_reason": match_reason,
        "check_url_score": check_url_score,
        "check_url_pass": check_url_pass,
        "check_link": check_link,
        "action_note": action_note,
    }


def process_single_row_no_chrome(
    row_data,
    master_name_col,
    child_name_col,
    master_address_col,
    child_address_col,
    ota_link_col,
    has_master_child_layout,
    url_name_threshold,
):
    index, row = row_data
    master_name = normalize_text(row[master_name_col])
    child_name = normalize_text(row[child_name_col])
    master_address = normalize_text(row[master_address_col])
    child_address = normalize_text(row[child_address_col])

    name_exact_match = bool(master_name and child_name and master_name == child_name)
    address_match, address_match_mode, address_match_score = addresses_semantically_match(master_address, child_address)

    if has_master_child_layout and name_exact_match and address_match:
        return {
            "index": index,
            "matched": True,
            "found_name": child_name,
            "found_address": child_address,
            "name_score": 100,
            "address_score": address_match_score,
            "match_reason": "master_child_precheck_match",
            "check_url_score": 0,
            "check_url_pass": False,
            "check_link": False,
            "action_note": f"Match pre-check: tên + địa chỉ master/child ({address_match_mode})",
        }

    expected_name = child_name
    ota_url = str(row[ota_link_col]).strip()
    if ota_url.startswith("http"):
        inferred_name = infer_name_from_url(ota_url)
        url_name_score = compute_name_score(expected_name, inferred_name) if inferred_name else 0
        check_url_pass = url_name_score >= url_name_threshold
        if check_url_pass:
            return {
                "index": index,
                "matched": True,
                "found_name": inferred_name,
                "found_address": "",
                "name_score": url_name_score,
                "address_score": 0,
                "match_reason": "url_name_high_confidence",
                "check_url_score": url_name_score,
                "check_url_pass": True,
                "check_link": False,
                "action_note": "Match: tên khớp cao từ URL (không Chrome)",
            }

        return {
            "index": index,
            "matched": False,
            "found_name": inferred_name,
            "found_address": "",
            "name_score": url_name_score,
            "address_score": 0,
            "match_reason": "url_name_low_confidence",
            "check_url_score": url_name_score,
            "check_url_pass": False,
            "check_link": False,
            "action_note": "Check URL không Chrome: độ tin cậy thấp",
        }

    return {
        "index": index,
        "matched": False,
        "found_name": "",
        "found_address": "",
        "name_score": 0,
        "address_score": 0,
        "match_reason": "invalid_url",
        "check_url_score": 0,
        "check_url_pass": False,
        "check_link": False,
        "action_note": "Bỏ qua: URL không hợp lệ",
    }


def append_case12_chain_vho_note(df, input_cols, match_reasons):
    total_rows = len(df)
    chain_list, chain_subbrand_pairs, chain_alias_groups = load_chain_subbrand_data("DanhSachChainBranch.xlsx")
    vho_terms = load_vho_terms("DanhSachVHO.xlsx")
    case12_chain_branch_vho_notes = []

    for index in range(total_rows):
        if bool(df.at[index, "matched trường hợp 1-2"]):
            if match_reasons[index] == "master_child_precheck_match":
                case12_chain_branch_vho_notes.append("Trường hợp 1-2: tên và địa chỉ giống nhau")
            else:
                if bool(df.at[index, "check_url_pass"]):
                    case12_chain_branch_vho_notes.append("Tên child nằm trên URL link child")
                else:
                    case12_chain_branch_vho_notes.append("")
            continue

        if bool(df.at[index, "check_url_pass"]):
            case12_chain_branch_vho_notes.append("Tên child nằm trên URL link child")
            continue

        master_name = normalize_text(df.at[index, input_cols["master_name_col"]])
        child_name = normalize_text(df.at[index, input_cols["child_name_col"]])
        chain_branch_note = classify_chain_branch_case(
            master_name,
            child_name,
            chain_list,
            chain_subbrand_pairs,
            chain_alias_groups,
            vho_terms,
        )
        case12_chain_branch_vho_notes.append(chain_branch_note)

    df["Case1-2_chain_branch_vho_CheckChildVoiURLCuaChild_note"] = case12_chain_branch_vho_notes


def verify_hotels_file_case12_chain_vho_no_chrome(
    input_path: str,
    output_path: str,
    url_name_threshold: int = 85,
    progress_callback=None,
):
    df = pd.read_excel(input_path)
    input_cols = get_input_columns(df)

    results = {}
    completed_count = 0
    total_rows = len(df)
    for index, row in df.iterrows():
        result = process_single_row_no_chrome(
            (index, row),
            input_cols["master_name_col"],
            input_cols["child_name_col"],
            input_cols["master_address_col"],
            input_cols["child_address_col"],
            input_cols["ota_link_col"],
            input_cols["has_master_child_layout"],
            url_name_threshold,
        )
        results[result["index"]] = result
        completed_count += 1
        if progress_callback:
            progress_callback(completed_count, total_rows)

    matched_values = []
    found_names = []
    found_addresses = []
    name_scores = []
    address_scores = []
    match_reasons = []
    check_url_scores = []
    check_url_passes = []
    check_links = []
    action_notes = []

    for index in range(total_rows):
        result = results.get(index, {
            "matched": False,
            "found_name": "",
            "found_address": "",
            "name_score": 0,
            "address_score": 0,
            "match_reason": "processing_failed",
            "check_url_score": 0,
            "check_url_pass": False,
            "check_link": False,
            "action_note": "Lỗi: không xử lý được dòng",
        })
        matched_values.append(result["matched"])
        found_names.append(result["found_name"])
        found_addresses.append(result["found_address"])
        name_scores.append(result["name_score"])
        address_scores.append(result["address_score"])
        match_reasons.append(result["match_reason"])
        check_url_scores.append(result["check_url_score"])
        check_url_passes.append(result["check_url_pass"])
        check_links.append(result["check_link"])
        action_notes.append(result["action_note"])

    df["matched trường hợp 1-2"] = matched_values
    df["found_name"] = found_names
    df["found_address"] = found_addresses
    df["name_score"] = name_scores
    df["address_score"] = address_scores
    df["match_reason"] = match_reasons
    df["check_url_score"] = check_url_scores
    df["check_url_pass"] = check_url_passes
    df["check_link"] = check_links
    df["action_note"] = action_notes

    append_case12_chain_vho_note(df, input_cols, match_reasons)

    df.to_excel(output_path, index=False)
    highlight_column_yellow(output_path, "Case1-2_chain_branch_vho_CheckChildVoiURLCuaChild_note")
    return output_path


def verify_hotels_file_ota_chrome(
    input_path: str,
    output_path: str,
    name_threshold: int = 75,
    address_threshold: int = 60,
    strong_name_threshold: int = 95,
    url_name_threshold: int = 85,
    headless: bool = False,
    num_workers: int = 3,
    progress_callback=None,
):
    return verify_hotels_file(
        input_path=input_path,
        output_path=output_path,
        name_threshold=name_threshold,
        address_threshold=address_threshold,
        strong_name_threshold=strong_name_threshold,
        url_name_threshold=url_name_threshold,
        headless=headless,
        num_workers=num_workers,
        progress_callback=progress_callback,
    )


def verify_hotels_file(
    input_path: str,
    output_path: str,
    name_threshold: int = 75,
    address_threshold: int = 60,
    strong_name_threshold: int = 95,
    url_name_threshold: int = 85,
    headless: bool = False,
    num_workers: int = 3,
    progress_callback=None,
):
    return verify_hotels_file_ota_chrome_full(
        input_path=input_path,
        output_path=output_path,
        name_threshold=name_threshold,
        address_threshold=address_threshold,
        strong_name_threshold=strong_name_threshold,
        url_name_threshold=url_name_threshold,
        headless=headless,
        num_workers=num_workers,
        progress_callback=progress_callback,
    )


def verify_hotels_file_ota_chrome_full(
    input_path: str,
    output_path: str,
    name_threshold: int = 75,
    address_threshold: int = 60,
    strong_name_threshold: int = 95,
    url_name_threshold: int = 85,
    headless: bool = False,
    num_workers: int = 3,
    progress_callback=None,
):
    df = pd.read_excel(input_path)
    input_cols = get_input_columns(df)

    page_cache = {}
    cache_lock = Lock()
    results = {}
    completed_count = 0
    total_rows = len(df)

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {
            executor.submit(
                process_single_row,
                (index, row),
                input_cols["master_name_col"],
                input_cols["child_name_col"],
                input_cols["master_address_col"],
                input_cols["child_address_col"],
                input_cols["ota_link_col"],
                input_cols["has_master_child_layout"],
                name_threshold,
                address_threshold,
                strong_name_threshold,
                url_name_threshold,
                headless,
                page_cache,
                cache_lock,
            ): index
            for index, row in df.iterrows()
        }

        for future in as_completed(futures):
            try:
                result = future.result()
                results[result["index"]] = result
                completed_count += 1
                if progress_callback:
                    progress_callback(completed_count, total_rows)
            except Exception:
                pass

    matched_values = []
    found_names = []
    found_addresses = []
    name_scores = []
    address_scores = []
    match_reasons = []
    check_url_scores = []
    check_url_passes = []
    check_links = []
    action_notes = []

    for index in range(total_rows):
        result = results.get(index, {
            "matched": False,
            "found_name": "",
            "found_address": "",
            "name_score": 0,
            "address_score": 0,
            "match_reason": "processing_failed",
            "check_url_score": 0,
            "check_url_pass": False,
            "check_link": False,
            "action_note": "Lỗi: không xử lý được dòng",
        })
        matched_values.append(result["matched"])
        found_names.append(result["found_name"])
        found_addresses.append(result["found_address"])
        name_scores.append(result["name_score"])
        address_scores.append(result["address_score"])
        match_reasons.append(result["match_reason"])
        check_url_scores.append(result["check_url_score"])
        check_url_passes.append(result["check_url_pass"])
        check_links.append(result["check_link"])
        action_notes.append(result["action_note"])

    df["matched trường hợp 1-2"] = matched_values
    df["found_name"] = found_names
    df["found_address"] = found_addresses
    df["name_score"] = name_scores
    df["address_score"] = address_scores
    df["match_reason"] = match_reasons
    df["check_url_score"] = check_url_scores
    df["check_url_pass"] = check_url_passes
    df["check_link"] = check_links
    df["action_note"] = action_notes

    append_case12_chain_vho_note(df, input_cols, match_reasons)

    df.to_excel(output_path, index=False)
    highlight_column_yellow(output_path, "Case1-2_chain_branch_vho_CheckChildVoiURLCuaChild_note")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Xác minh tên + địa chỉ khách sạn từ link OTA")
    parser.add_argument("--input", required=True, help="Đường dẫn file Excel đầu vào")
    parser.add_argument("--output", default="verified_hotels.xlsx", help="Đường dẫn file Excel đầu ra")
    parser.add_argument("--mode", choices=["case12_no_chrome", "ota_chrome"], default="ota_chrome", help="Chế độ chạy")
    parser.add_argument("--name-threshold", type=int, default=75, help="Ngưỡng điểm tên (0-100)")
    parser.add_argument("--address-threshold", type=int, default=60, help="Ngưỡng điểm địa chỉ (0-100)")
    parser.add_argument("--url-name-threshold", type=int, default=85, help="Nếu điểm tên khớp từ URL >= ngưỡng này thì bỏ qua crawl")
    parser.add_argument("--workers", type=int, default=3, help="Số lượng browser chạy song song")
    parser.add_argument("--headless", action="store_true", help="Chạy Chrome ẩn")
    args = parser.parse_args()

    if args.mode == "case12_no_chrome":
        output_path = verify_hotels_file_case12_chain_vho_no_chrome(
            input_path=args.input,
            output_path=args.output,
            url_name_threshold=args.url_name_threshold,
        )
    else:
        output_path = verify_hotels_file_ota_chrome(
            input_path=args.input,
            output_path=args.output,
            name_threshold=args.name_threshold,
            address_threshold=args.address_threshold,
            url_name_threshold=args.url_name_threshold,
            num_workers=args.workers,
            headless=args.headless,
        )
    print(f"Đã hoàn tất. File kết quả: {output_path}")


if __name__ == "__main__":
    main()
